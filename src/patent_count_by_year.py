"""
patent_count_by_year.py
=======================
Annual counts of patent applications and DOCDB patent families for three
datasets:
  - Green patents       (green_patent8526.parquet)
  - High-influence green patents (high_influence_green_patents.parquet)
  - Neighbor patent families     (neighbor_families.parquet)

Year attribution
----------------
  Green / Hi  : ``earliest_filing_date`` (family-level earliest filing;
                 consistent with OECD practice and the quality-index pipeline).
  Neighbor     : ``year`` column (family year pre-computed in find_neighbor.py).

Outputs
-------
  PATSTAT2025FALL/output/vis/patent_count_by_year.xlsx  — multi-sheet workbook
  PATSTAT2025FALL/output/vis/FIG_count_by_year.png      — 4-panel bar chart

Run from the project root:
    python src/patent_count_by_year.py
"""

import os, warnings
import numpy as np
import polars as pl
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ── Style ─────────────────────────────────────────────────────────────────────
plt.style.use("seaborn-v0_8-paper")
plt.rcParams.update({
    "text.usetex":    True,
    "font.family":    "serif",
    "font.serif":     ["Times New Roman", "Computer Modern Roman"],
    "font.size":      10,
    "axes.titlesize": 13,
    "axes.labelsize": 10,
    "legend.fontsize": 8.5,
    "xtick.labelsize": 8,
    "ytick.labelsize": 8,
    "figure.dpi":     300,
    "axes.grid":      True,
    "grid.alpha":     0.35,
    "grid.linewidth": 0.5,
})

YEAR_MIN, YEAR_MAX = 1985, 2025
OUT = "PATSTAT2025FALL/output/vis"
os.makedirs(OUT, exist_ok=True)

# ── Colour palette ─────────────────────────────────────────────────────────────
C_GREEN_APP  = "#2ca02c"   # dark green  – green applications
C_GREEN_FAM  = "#98df8a"   # light green – green families
C_HI_APP     = "#d62728"   # dark red    – hi-influence applications
C_HI_FAM     = "#ff9896"   # light red   – hi-influence families
C_NB_FAM     = "#1f77b4"   # blue        – neighbor families
C_RATIO      = "#ff7f0e"   # orange      – hi/green ratio line


# ═══════════════════════════════════════════════════════════════════════════════
# 1.  Load & aggregate
# ═══════════════════════════════════════════════════════════════════════════════
print("Loading green_patent8526.parquet ...")
green = pl.read_parquet(
    "PATSTAT2025FALL/output/green_patent8526.parquet",
    columns=["appln_id", "docdb_family_id", "earliest_filing_date"],
).with_columns(
    pl.col("earliest_filing_date").str.slice(0, 4).cast(pl.Int32).alias("year")
).filter(
    (pl.col("year") >= YEAR_MIN) & (pl.col("year") <= YEAR_MAX)
)

print("Loading high_influence_green_patents.parquet ...")
hi = pl.read_parquet(
    "PATSTAT2025FALL/output/high_influence_green_patents.parquet",
    columns=["appln_id", "docdb_family_id", "earliest_filing_date"],
).with_columns(
    pl.col("earliest_filing_date").str.slice(0, 4).cast(pl.Int32).alias("year")
).filter(
    (pl.col("year") >= YEAR_MIN) & (pl.col("year") <= YEAR_MAX)
)

print("Loading neighbor_families.parquet ...")
nb_all = pl.read_parquet(
    "PATSTAT2025FALL/output/neighbor_families.parquet",
    columns=["docdb_family_id", "year"],
)
# Keep ALL neighbor families for total counts.
# 'year' can be pre-1985 or 9999 (unknown); we bin those separately.
nb = nb_all  # unfiltered; annual histogram uses 1985-2025 range only

years = list(range(YEAR_MIN, YEAR_MAX + 1))

# ── Green: applications and families per year ──────────────────────────────────
green_apps = (
    green.group_by("year")
    .agg(pl.len().alias("green_apps"))
    .sort("year")
)
green_fam = (
    green.group_by(["docdb_family_id", "year"])
    .agg()                              # deduplicate: one row per (family, year)
    .group_by("year")
    .agg(pl.len().alias("green_fam"))   # count distinct families per year
    .sort("year")
)

# ── Hi-influence: applications and families per year ───────────────────────────
hi_apps = (
    hi.group_by("year")
    .agg(pl.len().alias("hi_apps"))
    .sort("year")
)
hi_fam = (
    hi.group_by(["docdb_family_id", "year"])
    .agg()
    .group_by("year")
    .agg(pl.len().alias("hi_fam"))
    .sort("year")
)

# ── Neighbor: families per year (already family-level) ────────────────────────
# Classify each row: in-range (1985-2025), pre-1985, or unknown (year=9999)
nb_tagged = nb.with_columns(
    pl.when((pl.col("year") >= YEAR_MIN) & (pl.col("year") <= YEAR_MAX))
    .then(pl.col("year").cast(pl.Int32))
    .otherwise(None)          # pre-1985 or 9999 → excluded from annual bars
    .alias("year_in_range")
)

nb_fam = (
    nb_tagged.filter(pl.col("year_in_range").is_not_null())
    .group_by("year_in_range")
    .agg(pl.col("docdb_family_id").n_unique().alias("nb_fam"))
    .rename({"year_in_range": "year"})
    .sort("year")
)

nb_total         = nb.height                                       # all rows
nb_pre1985       = nb.filter(pl.col("year") < YEAR_MIN).height
nb_unknown       = nb.filter(pl.col("year") == 9999).height
nb_in_range      = nb.filter(
    (pl.col("year") >= YEAR_MIN) & (pl.col("year") <= YEAR_MAX)
).height

# ── Assemble full-year grid (fill zeros for missing years) ─────────────────────
base = pl.DataFrame({"year": years})

df = (
    base
    .join(green_apps, on="year", how="left")
    .join(green_fam,  on="year", how="left")
    .join(hi_apps,    on="year", how="left")
    .join(hi_fam,     on="year", how="left")
    .join(nb_fam,     on="year", how="left")
    .with_columns(pl.all().fill_null(0))
    .with_columns(
        # Ratio: hi families as share of green families (%)
        pl.when(pl.col("green_fam") > 0)
        .then((pl.col("hi_fam") / pl.col("green_fam") * 100).round(2))
        .otherwise(0.0)
        .alias("hi_rate_pct")
    )
    .to_pandas()
)

print(f"\nYear range in histogram  : {YEAR_MIN}–{YEAR_MAX}")
print(f"Total green applications : {df['green_apps'].sum():>12,}")
print(f"Total green families     : {df['green_fam'].sum():>12,}")
print(f"Total hi-infl applications: {df['hi_apps'].sum():>12,}")
print(f"Total hi-infl families   : {df['hi_fam'].sum():>12,}")
print(f"Neighbor families total  : {nb_total:>12,}")
print(f"  in 1985-2025           : {nb_in_range:>12,}")
print(f"  pre-1985               : {nb_pre1985:>12,}")
print(f"  year=9999 (unknown)    : {nb_unknown:>12,}")


# ═══════════════════════════════════════════════════════════════════════════════
# 2.  Excel export
# ═══════════════════════════════════════════════════════════════════════════════
XLSX_PATH = f"{OUT}/patent_count_by_year.xlsx"
print(f"\nWriting Excel to {XLSX_PATH} ...")

# ── helper: format a worksheet ────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
ALT_FILL      = PatternFill("solid", fgColor="D9E1F2")   # light blue
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
CENTER        = Alignment(horizontal="center", vertical="center")
THIN          = Side(style="thin", color="BFBFBF")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_sheet(ws, col_widths: list[int]):
    """Apply header style + alternating row fill + column widths."""
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.alignment = CENTER
            cell.border    = BORDER
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"


with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:

    # Sheet 1: full annual table
    sheet1 = df[[
        "year", "green_apps", "green_fam",
        "hi_apps", "hi_fam", "hi_rate_pct",
        "nb_fam",
    ]].rename(columns={
        "year":        "Year",
        "green_apps":  "Green Apps",
        "green_fam":   "Green Families",
        "hi_apps":     "Hi-Infl. Apps",
        "hi_fam":      "Hi-Infl. Families",
        "hi_rate_pct": "Hi/Green Rate (%)",
        "nb_fam":      "Neighbor Families (1985-2025)",
    })

    # Append totals + neighbor breakdown rows
    totals_row = pd.DataFrame([{
        "Year":                          "TOTAL (1985-2025)",
        "Green Apps":                    int(df["green_apps"].sum()),
        "Green Families":                int(df["green_fam"].sum()),
        "Hi-Infl. Apps":                 int(df["hi_apps"].sum()),
        "Hi-Infl. Families":             int(df["hi_fam"].sum()),
        "Hi/Green Rate (%)":             round(df["hi_fam"].sum() / df["green_fam"].sum() * 100, 2),
        "Neighbor Families (1985-2025)": nb_in_range,
    }])
    extra_rows = pd.DataFrame([
        {"Year": "Neighbor – pre-1985",      "Neighbor Families (1985-2025)": nb_pre1985},
        {"Year": "Neighbor – year unknown",  "Neighbor Families (1985-2025)": nb_unknown},
        {"Year": "Neighbor – ALL (total)",   "Neighbor Families (1985-2025)": nb_total},
    ])
    sheet1 = pd.concat([sheet1, totals_row, extra_rows], ignore_index=True)
    sheet1.to_excel(writer, sheet_name="Annual Counts", index=False)
    ws1 = writer.sheets["Annual Counts"]
    style_sheet(ws1, [22, 14, 16, 15, 17, 18, 28])

    # Sheet 2: cumulative sums
    sheet2 = sheet1.copy()
    for c in ["Green Apps","Green Families","Hi-Infl. Apps",
               "Hi-Infl. Families","Neighbor Families"]:
        sheet2[c] = sheet2[c].cumsum()
    sheet2["Hi/Green Rate (%)"] = (
        sheet2["Hi-Infl. Families"] / sheet2["Green Families"] * 100
    ).round(2)
    sheet2.to_excel(writer, sheet_name="Cumulative", index=False)
    ws2 = writer.sheets["Cumulative"]
    style_sheet(ws2, [8, 14, 16, 15, 17, 18, 18])

    # Sheet 3: 5-year period summary
    df_p = df.copy()
    df_p["period"] = pd.cut(
        df_p["year"],
        bins=[1984, 1989, 1994, 1999, 2004, 2009, 2014, 2019, 2025],
        labels=["1985-89","1990-94","1995-99","2000-04",
                "2005-09","2010-14","2015-19","2020-25"],
    )
    sheet3 = (df_p.groupby("period", observed=True)
              .agg(
                  green_apps=("green_apps","sum"),
                  green_fam =("green_fam", "sum"),
                  hi_apps   =("hi_apps",   "sum"),
                  hi_fam    =("hi_fam",    "sum"),
                  nb_fam    =("nb_fam",    "sum"),
              )
              .assign(hi_rate_pct=lambda d:
                      (d["hi_fam"]/d["green_fam"]*100).round(2))
              .reset_index()
              .rename(columns={
                  "period":       "Period",
                  "green_apps":   "Green Apps",
                  "green_fam":    "Green Families",
                  "hi_apps":      "Hi-Infl. Apps",
                  "hi_fam":       "Hi-Infl. Families",
                  "hi_rate_pct":  "Hi/Green Rate (%)",
                  "nb_fam":       "Neighbor Families",
              }))
    sheet3.to_excel(writer, sheet_name="5-Year Periods", index=False)
    ws3 = writer.sheets["5-Year Periods"]
    style_sheet(ws3, [12, 14, 16, 15, 17, 18, 18])

print(f"  Sheets written: 'Annual Counts', 'Cumulative', '5-Year Periods'")


# ═══════════════════════════════════════════════════════════════════════════════
# 3.  4-panel bar chart
# ═══════════════════════════════════════════════════════════════════════════════
print("\nDrawing figure ...")

yrs = df["year"].values
fig, axes = plt.subplots(4, 1, figsize=(14, 16),
                          gridspec_kw={"hspace": 0.45})

# ── Helper: integer y-axis formatter ─────────────────────────────────────────
def fmt_k(x, pos):
    if x >= 1_000_000:
        return f"{x/1_000_000:.1f}M"
    if x >= 1_000:
        return f"{x/1_000:.0f}k"
    return f"{int(x)}"

# shaded region for incomplete years
INCOMPLETE_FROM = 2021

def shade_incomplete(ax):
    ax.axvspan(INCOMPLETE_FROM - 0.5, YEAR_MAX + 0.5,
               color="grey", alpha=0.10, zorder=0,
               label=r"Incomplete (granted-patent lag)")


# ── Panel (a): Green patent applications and families ────────────────────────
ax = axes[0]
w = 0.4
ax.bar(yrs - w/2, df["green_apps"], width=w, color=C_GREEN_APP,
       label=r"Applications", zorder=3)
ax.bar(yrs + w/2, df["green_fam"],  width=w, color=C_GREEN_FAM,
       label=r"Families",      zorder=3)
shade_incomplete(ax)
ax.set_title(r"\textbf{(a)} Green Patent Applications and Families per Year")
ax.set_ylabel(r"Count")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_k))
ax.set_xlim(YEAR_MIN - 0.8, YEAR_MAX + 0.8)
ax.legend(loc="upper left", framealpha=0.8)

# ── Panel (b): Hi-influence applications and families ────────────────────────
ax = axes[1]
ax.bar(yrs - w/2, df["hi_apps"], width=w, color=C_HI_APP,
       label=r"Hi-Influence Applications", zorder=3)
ax.bar(yrs + w/2, df["hi_fam"],  width=w, color=C_HI_FAM,
       label=r"Hi-Influence Families",     zorder=3)
shade_incomplete(ax)
ax.set_title(r"\textbf{(b)} High-Influence Green Patent Applications and Families per Year")
ax.set_ylabel(r"Count")
ax.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_k))
ax.set_xlim(YEAR_MIN - 0.8, YEAR_MAX + 0.8)
ax.legend(loc="upper left", framealpha=0.8)

# ── Panel (c): Hi/Green rate + neighbor families (dual axis) ─────────────────
ax  = axes[2]
ax2 = ax.twinx()

ax.bar(yrs, df["nb_fam"], color=C_NB_FAM, alpha=0.75,
       label=r"Neighbor Families", zorder=3)
ax2.plot(yrs, df["hi_rate_pct"], color=C_RATIO, linewidth=1.6,
         marker="o", markersize=2.5, zorder=4,
         label=r"Hi-Influence Rate (\%)")
shade_incomplete(ax)
ax.set_title(r"\textbf{(c)} Neighbor Families per Year "
             r"and Hi-Influence Rate (Hi Families $\div$ Green Families)")
ax.set_ylabel(r"Neighbor Families", color=C_NB_FAM)
ax.yaxis.set_major_formatter(mticker.FuncFormatter(fmt_k))
ax.tick_params(axis="y", labelcolor=C_NB_FAM)
ax2.set_ylabel(r"Hi-Influence Rate (\%)", color=C_RATIO)
ax2.tick_params(axis="y", labelcolor=C_RATIO)
ax2.set_ylim(0, 100)
ax.set_xlim(YEAR_MIN - 0.8, YEAR_MAX + 0.8)

lines1, labels1 = ax.get_legend_handles_labels()
lines2, labels2 = ax2.get_legend_handles_labels()
ax.legend(lines1 + lines2, labels1 + labels2, loc="upper left", framealpha=0.8)

# ── Panel (d): All three families normalised to 1985 = 1 ────────────────────
ax = axes[3]
base_g  = df.loc[df["year"] == YEAR_MIN, "green_fam"].values[0]  or 1
base_hi = df.loc[df["year"] == YEAR_MIN, "hi_fam"].values[0]     or 1
base_nb = df.loc[df["year"] == YEAR_MIN, "nb_fam"].values[0]     or 1

ax.plot(yrs, df["green_fam"] / base_g,  color=C_GREEN_APP, linewidth=1.6,
        label=r"Green Families (index, 1985=1)")
ax.plot(yrs, df["hi_fam"]   / base_hi, color=C_HI_APP,    linewidth=1.6,
        linestyle="--", label=r"Hi-Influence Families (index)")
ax.plot(yrs, df["nb_fam"]   / base_nb, color=C_NB_FAM,    linewidth=1.6,
        linestyle="-.", label=r"Neighbor Families (index)")
ax.axhline(1, color="grey", linewidth=0.8, linestyle=":")
shade_incomplete(ax)
ax.set_title(r"\textbf{(d)} Growth Index --- All Family Types (1985 $= 1$)")
ax.set_ylabel(r"Index (1985 $= 1$)")
ax.set_xlim(YEAR_MIN - 0.8, YEAR_MAX + 0.8)
ax.legend(loc="upper left", framealpha=0.8)

# ── Shared x-axis formatting ──────────────────────────────────────────────────
for ax in axes:
    ax.set_xlabel(r"Earliest Filing Year")
    ax.set_xticks(range(YEAR_MIN, YEAR_MAX + 1, 5))
    ax.set_xticklabels([str(y) for y in range(YEAR_MIN, YEAR_MAX + 1, 5)],
                        rotation=45, ha="right")

fig.suptitle(
    r"\textbf{Annual Patent Counts by Type} "
    r"$\cdot$ PATSTAT 2025 Autumn $\cdot$ Earliest Filing Year 1985--2025",
    fontsize=12, y=1.003,
)
fig.text(
    0.99, 0.0,
    r"Grey shading: 2021--2025 counts are incomplete due to granted-patent processing lag. "
    r"Hi-Influence = top forward-citation tercile + family size / triadic criterion.",
    ha="right", fontsize=7, color="#555",
)

FIG_PATH = f"{OUT}/FIG_count_by_year.png"
fig.savefig(FIG_PATH, bbox_inches="tight", dpi=300)
plt.close()
print(f"Saved {FIG_PATH}")


# ═══════════════════════════════════════════════════════════════════════════════
# 4.  Console summary
# ═══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 75)
print(f"{'Year':>6}  {'GreenApp':>9}  {'GreenFam':>9}  "
      f"{'HiApp':>7}  {'HiFam':>7}  {'Hi%':>6}  {'NbFam':>10}")
print("-" * 75)
for _, r in df.iterrows():
    incomplete = "*" if r["year"] >= INCOMPLETE_FROM else " "
    print(f"{int(r['year']):>6}{incomplete} "
          f"{int(r['green_apps']):>9,}  {int(r['green_fam']):>9,}  "
          f"{int(r['hi_apps']):>7,}  {int(r['hi_fam']):>7,}  "
          f"{r['hi_rate_pct']:>5.1f}%  {int(r['nb_fam']):>10,}")
print("=" * 75)
print("  * = incomplete year (granted-patent processing lag)")
print(f"\nNeighbor family totals:")
print(f"  In 1985-2025  : {nb_in_range:>10,}")
print(f"  Pre-1985      : {nb_pre1985:>10,}  (cited by green patents but filed earlier)")
print(f"  Year unknown  : {nb_unknown:>10,}  (year=9999 in source data)")
print(f"  GRAND TOTAL   : {nb_total:>10,}")
print(f"\nExcel  -> {XLSX_PATH}")
print(f"Figure -> {FIG_PATH}")
